#!/usr/bin/env python3
"""
任务4：故障报警预测
- 加载训练好的 XGBoost 模型 (best_model.pkl)
- 分块读取附件2的 M201.csv / M202.csv（大文件 ~875MB）
- 预处理：与训练数据一致的特征工程
- 预测故障事件（连续故障记录合并为一个事件）
- 输出 result2.xlsx（与模板格式一致）
"""

from __future__ import annotations

import os
import sys
import time as _time
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
import joblib
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── 路径配置 ──────────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent.parent
MODEL_PATH = Path(__file__).resolve().parent / "output" / "task4_models" / "best_model.pkl"
OUTPUT_DIR = Path(__file__).resolve().parent / "output"
OUTPUT_FILE = OUTPUT_DIR / "result2.xlsx"

# 附件2 数据目录
ATTACH2_DIR = (
    PROJECT_ROOT
    / "课程设计资料"
    / "泰迪杯2024年A题_说明及数据"
    / "A题-示例数据"
    / "附件2"
)

# ── 传感器特征列（与训练一致）──────────────────────────────────────────
SENSOR_COLS = [
    "物料推送气缸推送状态", "物料推送气缸收回状态", "物料推送数", "物料待抓取数",
    "放置容器数", "容器上传检测数", "填装检测数", "填装定位器固定状态", "填装定位器放开状态",
    "物料抓取数", "填装旋转数", "填装下降数", "填装数", "加盖检测数", "加盖定位数",
    "推盖数", "加盖下降数", "加盖数", "拧盖检测数", "拧盖定位数", "拧盖下降数",
    "拧盖旋转数", "拧盖数", "合格数", "不合格数",
]

# ── 故障类型列 ─────────────────────────────────────────────────────
FAULT_COLS = [
    "物料推送装置故障1001", "物料检测装置故障2001", "填装装置检测故障4001",
    "填装装置定位故障4002", "填装装置填装故障4003", "加盖装置定位故障5001",
    "加盖装置加盖故障5002", "拧盖装置定位故障6001", "拧盖装置拧盖故障6002",
]

# 故障代码 → 缩短名称映射
FAULT_CODE_MAP = {
    "物料推送装置故障1001": 1001,
    "物料检测装置故障2001": 2001,
    "填装装置检测故障4001": 4001,
    "填装装置定位故障4002": 4002,
    "填装装置填装故障4003": 4003,
    "加盖装置定位故障5001": 5001,
    "加盖装置加盖故障5002": 5002,
    "拧盖装置定位故障6001": 6001,
    "拧盖装置拧盖故障6002": 6002,
}

# 故障代码 → 对应的关键传感器组（用于推断故障类型）
FAULT_SENSOR_MAP = {
    1001: ["物料推送气缸推送状态", "物料推送气缸收回状态", "物料推送数"],
    2001: ["物料待抓取数", "物料抓取数"],
    4001: ["填装检测数"],
    4002: ["填装定位器固定状态", "填装定位器放开状态"],
    4003: ["填装旋转数", "填装下降数", "填装数"],
    5001: ["加盖检测数", "加盖定位数"],
    5002: ["推盖数", "加盖下降数", "加盖数"],
    6001: ["拧盖检测数", "拧盖定位数"],
    6002: ["拧盖下降数", "拧盖旋转数", "拧盖数"],
}

# 分块大小
CHUNKSIZE = 50000

# 附件2编码（实测为 UTF-8）
ENCODING = "utf-8"


# ── 辅助函数 ──────────────────────────────────────────────────────

def preprocess_chunk(chunk: pd.DataFrame, line_name: str) -> pd.DataFrame:
    """
    对一个数据块做与训练一致的预处理。
    返回包含特征列的 DataFrame。
    """
    # 1. 数值转换（传感器列）
    for col in SENSOR_COLS:
        if col in chunk.columns:
            chunk[col] = pd.to_numeric(chunk[col], errors="coerce")

    # 故障列数值转换
    for col in FAULT_COLS:
        if col in chunk.columns:
            chunk[col] = pd.to_numeric(chunk[col], errors="coerce")

    # 2. 删除传感器全空行
    existing_sensor = [c for c in SENSOR_COLS if c in chunk.columns]
    chunk = chunk.dropna(subset=existing_sensor, how="all")

    # 3. 填充 NaN 为 0（传感器列）
    chunk[existing_sensor] = chunk[existing_sensor].fillna(0)

    # 4. 时间特征：附件2的"时间"列是秒数，需要转换为小时
    #    训练数据的 hour_of_day 是直接取时间列（0-7表示小时）
    #    附件2的 时间 是秒数 (0 ~ 29058)，需要 // 3600 得到小时
    chunk["hour_of_day"] = (chunk["时间"] // 3600).astype(int)
    chunk["is_daytime"] = chunk["hour_of_day"].between(6, 18).astype(int)

    # 5. one-hot 编码：模型训练时使用了 pd.get_dummies(df["生产线编号"], prefix="line")
    #    训练数据只有 M101，所以特征里只有 line_M101
    #    附件2 是 M201/M202，line_M101 应为 0
    chunk["line_M101"] = 0

    return chunk


def build_features(chunk: pd.DataFrame) -> pd.DataFrame:
    """从预处理后的 chunk 中提取模型所需的特征矩阵。"""
    feature_cols = SENSOR_COLS + ["hour_of_day", "is_daytime", "line_M101"]
    existing = [c for c in feature_cols if c in chunk.columns]
    return chunk[existing].copy()


def infer_fault_type(row: pd.Series) -> int:
    """
    推断单行的故障类型。
    策略：检查各故障类型对应的关键传感器组，
    如果某组传感器全为 0 且当前行被模型预测为故障，
    则认为该组对应的故障类型最可能。
    如果无法判断，返回最常见的故障类型 1001。
    """
    # 收集各故障类型的"异常分数"
    scores: dict[int, float] = {}
    for code, sensors in FAULT_SENSOR_MAP.items():
        vals = [row.get(s, 0) for s in sensors if s in row.index]
        if not vals:
            continue
        # 异常分数：传感器值的方差越大，越可能是正常运行后出故障
        # 如果传感器全为0，说明该环节完全停滞，是故障信号
        all_zero = all(v == 0 for v in vals)
        if all_zero:
            scores[code] = 1.0
        else:
            # 有值但异常：值偏低或偏高
            scores[code] = 0.3

    if not scores:
        return 1001  # 默认

    # 返回异常分数最高的故障类型
    return max(scores, key=lambda k: scores[k])


def merge_fault_events(
    fault_df: pd.DataFrame,
    line_name: str,
) -> list[dict]:
    """
    将预测为故障的行合并为故障事件。
    连续的故障行（同一天、连续时间）合并为一个事件。
    每个事件记录：生产线编号、日期、开始时间、持续时长(秒)、故障类型。
    """
    if fault_df.empty:
        return []

    fault_df = fault_df.sort_values(["日期", "时间"]).reset_index(drop=True)
    events: list[dict] = []

    i = 0
    while i < len(fault_df):
        row_i = fault_df.iloc[i]
        start_date = row_i["日期"]
        start_time = row_i["时间"]
        fault_type = int(row_i.get("_fault_type", 1001))

        # 向后查找连续的故障行（同一天，时间差 ≤ 2秒视为连续）
        j = i + 1
        while j < len(fault_df):
            row_j = fault_df.iloc[j]
            # 如果是同一天且时间连续（差值≤2秒）
            if (
                row_j["日期"] == start_date
                and (row_j["时间"] - fault_df.iloc[j - 1]["时间"]) <= 2
            ):
                j += 1
            else:
                break

        end_time = fault_df.iloc[j - 1]["时间"]
        duration = int(end_time - start_time + 1)  # 持续时长（秒）

        events.append({
            "生产线编号": line_name,
            "日期": int(start_date),
            "开始时间": int(start_time),
            "持续时长_秒": duration,
            "故障类型": fault_type,
        })

        i = j

    return events


def process_line(
    line_name: str,
    model_data: dict,
) -> list[dict]:
    """
    处理一条生产线的数据（M201 或 M202）。
    分块读取、预处理、预测、合并故障事件。
    """
    csv_path = ATTACH2_DIR / f"{line_name}.csv"
    if not csv_path.exists():
        print(f"  ⚠ 文件不存在: {csv_path}")
        return []

    print(f"\n{'='*60}")
    print(f"处理 {line_name}: {csv_path}")
    print(f"文件大小: {csv_path.stat().st_size / 1024 / 1024:.0f} MB")
    print(f"{'='*60}")

    model = model_data["model"]
    feature_names = model_data["feature_names"]
    all_events: list[dict] = []

    t_start = _time.time()
    chunk_idx = 0
    total_rows = 0
    total_fault_rows = 0

    for chunk in pd.read_csv(
        csv_path,
        encoding=ENCODING,
        chunksize=CHUNKSIZE,
        low_memory=False,
    ):
        chunk_idx += 1
        total_rows += len(chunk)

        # 预处理
        chunk = preprocess_chunk(chunk, line_name)
        if chunk.empty:
            continue

        # 构建特征矩阵
        X = build_features(chunk)

        # 对齐特征列（确保与训练时一致）
        for col in feature_names:
            if col not in X.columns:
                X[col] = 0
        X = X[feature_names]

        # 预测
        y_pred = model.predict(X)
        chunk["_pred"] = y_pred

        # 筛选预测为故障的行
        fault_mask = chunk["_pred"] == 1
        fault_chunk = chunk[fault_mask].copy()
        n_fault = len(fault_chunk)
        total_fault_rows += n_fault

        # 对故障行推断故障类型
        if n_fault > 0:
            fault_chunk["_fault_type"] = fault_chunk.apply(infer_fault_type, axis=1)
            # 合并为故障事件
            events = merge_fault_events(fault_chunk, line_name)
            all_events.extend(events)

        # 打印进度
        elapsed = _time.time() - t_start
        print(
            f"  块 {chunk_idx:>4d} | "
            f"已处理 {total_rows:>10,} 行 | "
            f"故障行 {total_fault_rows:>8,} | "
            f"故障事件 {len(all_events):>6d} | "
            f"耗时 {elapsed:.1f}s"
        )

    elapsed_total = _time.time() - t_start
    print(f"\n  {line_name} 完成: 共 {total_rows:,} 行, "
          f"预测故障行 {total_fault_rows:,}, "
          f"故障事件 {len(all_events)} 个, "
          f"耗时 {elapsed_total:.1f}s")

    return all_events


def events_to_excel(all_events: dict[str, list[dict]], output_path: Path) -> None:
    """
    将故障事件写入 Excel，格式与模板一致。
    每条生产线一个 sheet，每种故障类型 3 列（日期、开始时间、持续时长/秒）。
    """
    wb = Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    fault_codes = [1001, 2001, 4001, 4002, 4003, 5001, 5002, 6001, 6002]

    for line_name in ["M201", "M202"]:
        ws = wb.create_sheet(title=line_name)
        events = all_events.get(line_name, [])

        # ── 第1行：故障代码表头（合并单元格）──
        ws.cell(row=1, column=1, value="故障编号")
        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        col = 2
        for code in fault_codes:
            ws.cell(row=1, column=col, value=code)
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")
            # 合并 3 列（日期、开始时间、持续时长）
            ws.merge_cells(
                start_row=1, start_column=col,
                end_row=1, end_column=col + 2,
            )
            col += 3

        # ── 第2行：子表头 ──
        ws.cell(row=2, column=1, value="序号")
        ws.cell(row=2, column=1).font = Font(bold=True)
        col = 2
        for _ in fault_codes:
            ws.cell(row=2, column=col, value="日期")
            ws.cell(row=2, column=col + 1, value="开始时间")
            ws.cell(row=2, column=col + 2, value="持续时长/秒")
            for c in range(col, col + 3):
                ws.cell(row=2, column=c).font = Font(bold=True)
                ws.cell(row=2, column=c).alignment = Alignment(horizontal="center")
            col += 3

        # ── 数据行：按故障类型分组 ──
        # 将事件按故障类型分组
        events_by_type: dict[int, list[dict]] = {code: [] for code in fault_codes}
        for ev in events:
            ftype = ev["故障类型"]
            if ftype in events_by_type:
                events_by_type[ftype].append(ev)

        # 找到最大行数
        max_rows = max((len(v) for v in events_by_type.values()), default=0)

        if max_rows == 0:
            # 无故障事件，写一行空行
            ws.cell(row=3, column=1, value=1)
        else:
            for row_idx in range(max_rows):
                excel_row = row_idx + 3
                ws.cell(row=excel_row, column=1, value=row_idx + 1)

                col = 2
                for code in fault_codes:
                    evts = events_by_type[code]
                    if row_idx < len(evts):
                        ev = evts[row_idx]
                        ws.cell(row=excel_row, column=col, value=ev["日期"])
                        ws.cell(row=excel_row, column=col + 1, value=ev["开始时间"])
                        ws.cell(row=excel_row, column=col + 2, value=ev["持续时长_秒"])
                    col += 3

        # 设置列宽
        ws.column_dimensions["A"].width = 8
        col_letter_idx = 2
        for _ in fault_codes:
            for offset in range(3):
                letter = get_column_letter(col_letter_idx + offset)
                ws.column_dimensions[letter].width = 14
            col_letter_idx += 3

    wb.save(output_path)
    print(f"\n✓ 结果已保存: {output_path}")


def main() -> None:
    print("=" * 60)
    print("任务4：故障报警预测 — 附件2 (M201/M202)")
    print("=" * 60)

    # ── 1. 加载模型 ──────────────────────────────────────────────
    print(f"\n[1/3] 加载模型: {MODEL_PATH}")
    if not MODEL_PATH.exists():
        print(f"  ✗ 模型文件不存在: {MODEL_PATH}")
        print("  请先运行 task4_model.py 训练并保存模型。")
        sys.exit(1)

    model_data = joblib.load(MODEL_PATH)
    print(f"  模型类型: {model_data.get('model_name', 'unknown')}")
    print(f"  训练 AUC: {model_data.get('auc', 'N/A')}")
    print(f"  特征数量: {len(model_data.get('feature_names', []))}")

    # ── 2. 逐线处理 ──────────────────────────────────────────────
    print("\n[2/3] 处理生产线数据...")
    all_events: dict[str, list[dict]] = {}

    for line_name in ["M201", "M202"]:
        events = process_line(line_name, model_data)
        all_events[line_name] = events

    # ── 3. 输出 Excel ────────────────────────────────────────────
    print(f"\n[3/3] 生成结果文件...")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    events_to_excel(all_events, OUTPUT_FILE)

    # ── 汇总 ──────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("预测完成！汇总:")
    print("=" * 60)
    for line_name in ["M201", "M202"]:
        events = all_events.get(line_name, [])
        print(f"  {line_name}: {len(events)} 个故障事件")
        # 按故障类型统计
        type_counts: dict[int, int] = {}
        for ev in events:
            t = ev["故障类型"]
            type_counts[t] = type_counts.get(t, 0) + 1
        for code in sorted(type_counts.keys()):
            print(f"    故障 {code}: {type_counts[code]} 次")
    print(f"\n输出文件: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
